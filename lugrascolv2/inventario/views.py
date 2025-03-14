from datetime import datetime
import json,io
import random
from django.forms import DecimalField
from django.http import HttpResponse, HttpResponseBadRequest, JsonResponse,FileResponse
from django.shortcuts import get_object_or_404, render, redirect
from facturacion.models import  Inventario, TransaccionAjuste, Ajustes, Averias, TransaccionAverias, TransMp, Proveedores, Transformulas, TransaccionFactura, TransaccionRemision, TransaccionRemision, SalidasMpOrden, Compras, TransaccionOrden
from django.db.models import Max, F, IntegerField, Sum
from django.db.models.functions import Coalesce
from docx import Document # type: ignore
from reportlab.lib.pagesizes import letter # type: ignore
from reportlab.platypus import SimpleDocTemplate, Paragraph, Table, TableStyle, Spacer # type: ignore
from reportlab.lib.styles import getSampleStyleSheet # type: ignore
from reportlab.lib import colors # type: ignore
from decimal import Decimal
from .forms import ExcelUploadForm
import pandas as pd # type: ignore
from django.db import IntegrityError, OperationalError
from django.utils.dateparse import parse_date
from openpyxl import Workbook # type: ignore
from openpyxl.utils import get_column_letter # type: ignore
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side # type: ignore

import logging


#Create your views here.

# Create your views here.


def inventory_view(request):
    inventario = Inventario.objects.all()
    return render(request, 'inventario.html', {"inventario" : inventario} )



def ajuste(request):
    producto = Inventario.objects.all()
    return render(request, 'ajuste_inventario.html', {'productos': producto})

def agregar_transaccion_ajuste(request):
        if request.method == 'POST':
            
            try:
                datos_tabla = json.loads(request.body.decode('utf-8')).get('datos_tabla')
                if datos_tabla:
                    for datos in datos_tabla:
                        codigo = datos['codigo']
                        descripcion = datos['descripcion']
                        fecha_ajuste = datos['fecha']
                        producto = datos['id_producto']
                        cantidad = datos['cantidad']
                    
                        
                        ajusteN, _ = Ajustes.objects.get_or_create(id_ajuste = codigo)
                        inventario, _ = Inventario.objects.get_or_create(cod_inventario= producto)
                        nueva_cantidad = F('cantidad') + cantidad
                        
                        nuevo_ajuste = TransaccionAjuste.objects.create(
                        fecha_ajuste = fecha_ajuste,
                        cant_ajuste = cantidad,
                        descripcion = descripcion,
                        id_ajuste = ajusteN,
                        cod_inventario = inventario
                        )
                        nuevo_ajuste.save();
                        # Aquí puedes procesar y guardar los datos adicionales si es necesario
                        Inventario.objects.filter(cod_inventario=producto).update(cantidad=nueva_cantidad)
                        
                return JsonResponse({'success': True, 'message': 'Datos actualizados correctamente'})
            except Exception as e:
                return JsonResponse({'success': False, 'error': str(e)}, status=400)
        else:
            return JsonResponse({'success': False, 'error': 'Método de solicitud no permitido'}, status=405)
        
        
        

def obtener_numero_ajuste(request):
    ultimo_ajuste = None  # Define la variable antes del bloque try
    
    try:
        ultimo_ajuste = Ajustes.objects.aggregate(Max('id_ajuste'))['id_ajuste__max']
        print(ultimo_ajuste)
        if ultimo_ajuste is not None:
            nuevo_ajuste = ultimo_ajuste + 1
        else:
            nuevo_ajuste = 10001
            print(nuevo_ajuste)
    except Exception as e:
        print(f"Error al obtener o generar el número de ajuste: {e}")
        nuevo_ajuste = None

    data = {'numero_ajuste': nuevo_ajuste, 'ultimo_ajuste': ultimo_ajuste}
    return JsonResponse(data)


def ver_ajuste (request):
    transaccion = TransaccionAjuste.objects.all()
    return  render(request, 'ajuste_inventario.html',{'transacciones': transaccion})



def agregar_transaccion_averias(request):
        if request.method == 'POST':
            
            try:
                datos_tabla = json.loads(request.body.decode('utf-8')).get('datos_tabla')
                if datos_tabla:
                    # Generar el documento de Word
                    document = Document()
                    document.add_heading('Transacciones de Averías', level=1)
                    for datos in datos_tabla:
                        codigo = datos['codigo']
                        fecha_ajuste = datos['fecha']
                        producto = datos['id_producto']
                        cantidad = datos['cantidad']
                        motivo = datos['motivo']
                        
                    
                        
                        averiaN, _ = Averias.objects.get_or_create(id_averia = codigo)
                        inventario, _ = Inventario.objects.get_or_create(cod_inventario= producto)
                        nueva_cantidad = F('cantidad') - cantidad
                        
                        nuevo_averia = TransaccionAverias.objects.create(
                        fecha_averia = fecha_ajuste,
                        cant_averia = cantidad,
                        id_averia = averiaN,
                        cod_inventario = inventario,
                        motivo = motivo
                        )
                        nuevo_averia.save();
                        # Aquí puedes procesar y guardar los datos adicionales si es necesario
                        Inventario.objects.filter(cod_inventario=producto).update(cantidad=nueva_cantidad)
                    
                return JsonResponse({'success': True, 'message': 'Datos actualizados correctamente'})
            except Exception as e:
                return JsonResponse({'success': False, 'error': str(e)}, status=400)
        else:
            return JsonResponse({'success': False, 'error': 'Método de solicitud no permitido'}, status=405)
        
        
        
def obtener_numero_averia(request):
    ultima_averia = None  # Define la variable antes del bloque try
    
    try:
        ultima_averia = Averias.objects.aggregate(Max('id_averia'))['id_averia__max']
        print(ultima_averia)
        if ultima_averia is not None:
            nueva_averia = ultima_averia + 1
        else:
            nueva_averia = 20001
            print(nueva_averia)
    except Exception as e:
        print(f"Error al obtener o generar el número de ajuste: {e}")
        nueva_averia = None

    data = {'numero_ajuste': nueva_averia, 'ultimo_ajuste': ultima_averia}
    return JsonResponse(data)

def averias(request):
    producto = Inventario.objects.all()
    return render(request, 'control_averias.html', {'productos': producto})



def generar_informe_averia(request):
    
    if request.method == 'GET':
        id_averia = request.GET.get('id_averia')
    elif request.method == 'POST':
        id_averia = request.POST.get('id_averia')
    else:
        # En este ejemplo, no manejamos otras solicitudes HTTP, como PUT o DELETE
        return HttpResponseBadRequest("Método HTTP no permitido")

    # Obtener las transacciones de averías asociadas a la avería
    transacciones = TransaccionAverias.objects.filter(id_averia=id_averia)

    # Obtener la fecha de la avería (suponiendo que todas las transacciones tienen la misma fecha)
    primera_transaccion = transacciones.first()  # Tomamos la primera transacción como representativa
    fecha_averia = primera_transaccion.fecha_averia if primera_transaccion else "Fecha no disponible"

    # Crear un documento PDF
    pdf_buffer = io.BytesIO()
    pdf = SimpleDocTemplate(pdf_buffer, pagesize=letter)

    # Estilos para el documento
    styles = getSampleStyleSheet()

    # Contenido del documento
    content = []

    # Título y fecha de la avería
    content.append(Paragraph(f"Informe de Avería #{id_averia}", styles['Title']))
    content.append(Spacer(1, 30)) 

    content.append(Paragraph(f"Fecha de Avería: {fecha_averia}", styles['Normal']))
    
    content.append(Spacer(1, 30))  

    # Detalles de las transacciones en forma de tabla
    data = [['Código de Inventario', 'Nombre', 'Cantidad', 'Costo']]
    totalCosto = Decimal('0.00')
    
    for transaccion in transacciones:
        
        formula = Transformulas.objects.filter(
                cod_inventario=transaccion.cod_inventario
        ).first()
            
        if formula:
            # Si el código del producto es una fórmula, calcular el costo usando la fórmula
            subtotal_costo = 0.0
            iva_costo = 0.0
            total_costo = 0.0
            
            # Calcular el subtotal_costo usando las materias primas de la fórmula
            for i in range(1, 11):  # Iterar sobre los campos materia1 hasta materia8
                campo_materia = getattr(formula, f"materia{i}")
                if campo_materia:  # Verificar si hay un código de materia prima en el campo actual
                    materia_prima = TransMp.objects.filter(cod_inventario=campo_materia).order_by('-fecha_ingreso').first()
                    if materia_prima:
                        setattr(formula, f"costo_unitario{i}", materia_prima.costo_unitario)
                        cantidad = getattr(formula, f"cant_materia{i}")
                        costo_unitario = getattr(materia_prima, f"costo_unitario")
                        subtotal_costo += cantidad * costo_unitario
                    else:
                        # Asignar valores predeterminados si la materia prima no se encuentra
                        setattr(formula, f"nombre_mp{i}", "Materia prima no encontrada")
                        setattr(formula, f"costo_unitario{i}", 0.0)
            
            iva_costo = (Decimal(subtotal_costo) * Decimal(formula.porcentajeiva)) / Decimal('100')
            total_costo = Decimal(subtotal_costo) + Decimal(formula.costosindirectos) + iva_costo
            precio_unitario = total_costo
            
        else:
            # Si el código del producto es una materia prima, obtener el costo unitario de la última transacción
            ultimo_transaccion = TransMp.objects.filter(
                cod_inventario=transaccion.cod_inventario
            ).order_by('-fecha_ingreso').first()
            
            if ultimo_transaccion:
                # Si existe una transacción, usa su costo unitario
                precio_unitario = Decimal(ultimo_transaccion.costo_unitario)
            else:
                # Si no hay transacción, establecer un precio predeterminado
                precio_unitario = Decimal('0.00')
                
        total_fila = Decimal(transaccion.cant_averia) * precio_unitario
        totalCosto += total_fila
        
        data.append([   transaccion.cod_inventario.cod_inventario,
                        transaccion.cod_inventario.nombre,
                        transaccion.cant_averia,
                        f"{precio_unitario:,.2f}"])
        
    
    # Crear la tabla
    table = Table(data)
    table.setStyle(TableStyle([ ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                                ('WORDWRAP', (0, 0), (-1, -1), True)]))

    content.append(table)
    content.append(Spacer(1, 20))
    content.append(Paragraph(f"Total Costo Averías: {totalCosto:,.2f}", styles['Normal']))

    # Generar el PDF
    pdf.build(content)

    # Devolver el archivo como una respuesta HTTP
    pdf_buffer.seek(0)
    response = HttpResponse(pdf_buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename=informe_averia_{id_averia}.pdf'
    return response



def buscarKardex(request):
    productos = Inventario.objects.all()
    return render(request, 'kardex.html', {'productos': productos})

def kardex_view(request):
    if request.method == 'POST':
        # Obtener datos del POST
        cod_producto = request.POST.get('cod_producto')
        fecha_inicio = request.POST.get('fecha_inicio')
        fecha_final = request.POST.get('fecha_final')

        # Filtrar transacciones según el rango de fechas y el producto
        transacciones = {
            'materia_prima': TransMp.objects.filter(
                cod_inventario=cod_producto,
                fecha_ingreso__range=[fecha_inicio,fecha_final]
            ),
            'ajuste': TransaccionAjuste.objects.filter(
                cod_inventario=cod_producto,
                fecha_ajuste__range=[fecha_inicio, fecha_final]
            ),
            'averia': TransaccionAverias.objects.filter(
                cod_inventario=cod_producto,
                fecha_averia__range=[fecha_inicio, fecha_final]
            ),
            'salida_orden': SalidasMpOrden.objects.filter(
                cod_inventario=cod_producto,
                fecha_e_produccion__range=[fecha_inicio, fecha_final]
            ),
            'entrada_produccion': TransaccionOrden.objects.filter(
                cod_inventario = cod_producto,
                fecha_terminacion_orden__range=[fecha_inicio, fecha_final]
            ),
            'factura': TransaccionFactura.objects.filter(
                cod_inventario=cod_producto,
                fecha_factura__range=[fecha_inicio, fecha_final]
            ),
            'remision': TransaccionRemision.objects.filter(
                cod_inventario=cod_producto,
                fecha_remision__range=[fecha_inicio, fecha_final]
            ),




        }

        # Inicializar el saldo de inventario
        saldo = 0



        # Crear una lista de resultados para el kardex
        kardex = []

        for key, qs in transacciones.items():
            for transaccion in qs:
                if key == 'factura':
                    fecha = transaccion.fecha_factura
                    cantidad = transaccion.cantidad
                    referencia = transaccion.nfactura.nfactura
                    entradas = 0
                    salidas = float(cantidad)
                elif key == 'remision':
                    fecha = transaccion.fecha_remision
                    cantidad = transaccion.cantidad
                    referencia = transaccion.nremision.nremision
                    entradas = 0
                    salidas = float(cantidad)
                elif key == 'ajuste':
                    fecha = transaccion.fecha_ajuste
                    cantidad = transaccion.cant_ajuste
                    referencia = transaccion.id_ajuste.id_ajuste
                    entradas = cantidad if cantidad > 0 else 0
                    salidas = -cantidad if cantidad < 0 else 0
                elif key == 'averia':
                    fecha = transaccion.fecha_averia
                    cantidad = transaccion.cant_averia
                    referencia = transaccion.id_averia.id_averia
                    entradas = 0
                    salidas = cantidad
                elif key == 'salida_orden':
                    fecha = transaccion.fecha_e_produccion
                    cantidad = transaccion.cantidad
                    referencia = transaccion.id_orden.id_orden
                    entradas = 0
                    salidas = float(cantidad)
                elif key == 'materia_prima':
                    fecha = transaccion.fecha_ingreso
                    cantidad = transaccion.cant_mp
                    referencia = transaccion.id_compra.id_compra
                    entradas = float(cantidad)
                    salidas = 0
                elif key == 'entrada_produccion':
                    fecha = transaccion.fecha_terminacion_orden
                    cantidad = transaccion.cantidad
                    referencia = transaccion.id_orden.id_orden
                    entradas = float(cantidad)
                    salidas = 0
                        
                

                saldo += entradas - salidas

                kardex.append({
                    'fecha': fecha.strftime('%Y-%m-%d'),
                    'descripcion': key,
                    'referencia': referencia,
                    'entradas': float(entradas),
                    'salidas': float(salidas),
                    'saldo': float(saldo),
                })
                
                print("Kardex:")
                for item in kardex:
                    print(item)


        # Ordenar el kardex por fecha
        kardex.sort(key=lambda x: x['fecha'])

        return JsonResponse({'kardex': kardex})
    else:
        return JsonResponse({'error': 'Método no permitido'}, status=405)
    
    
    
    
    
    
def cargaSaldos(request):
    return render(request, 'saldosIniciales.html')    



def upload_excel(request):
    if request.method == 'POST':
        excel_file = request.FILES.get('file')
        if not excel_file:
            return JsonResponse({'error': 'No se ha subido ningún archivo'}, status=400)

        try:
            df = pd.read_excel(excel_file, engine='openpyxl')
            
                        # Imprimir las primeras filas del dataframe para depuración
            
            # Convertir columnas específicas a solo fecha
            if 'fecha de ingreso (dd/mm/aaaa)' in df.columns:
                df['fecha de ingreso (dd/mm/aaaa)'] = pd.to_datetime(df['fecha de ingreso (dd/mm/aaaa)'], errors='coerce').dt.date

            data = df.to_dict(orient='records')
            
             # Imprimir los datos procesados para depuración
            
            
            
            
            return JsonResponse({'data': data})
        
            
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)

    return JsonResponse({'error': 'Método no permitido'}, status=405)

# Configurar el logger
logger = logging.getLogger(__name__)
logging.basicConfig(level=logging.DEBUG)
def guardarDatos(request):
    if request.method == 'POST':
        try:
            # Obtener los datos del cuerpo de la solicitud
            data = json.loads(request.body)
            
            # Obtener los datos enviados desde el frontend
            datos = data.get('datos', [])
            
            # Filtrar los datos por tipo
            datos_tipo_m = [item for item in datos if item.get('Tipo') == 'm']
            datos_tipo_pt = [item for item in datos if item.get('Tipo') == 'pt']


            
            # Procesar datos de tipo 'm'
            for item in datos_tipo_m:
                try:
                    # Imprimir los datos antes de convertir
                    print("Datos crudos del item:", item)
                    
                    # Convertir valores a los tipos correctos
                    codigo = int(item.get('Codigo'))
                    nombre = item.get('Nombre')
                    cantidad = int(item.get('Cantidad'))
                    costo_unitario = float(item.get('Costo Unitario', 0))
                    total_linea = float(item.get('Total', 0))         
                    fecha_ingreso= item.get('Fecha de ingreso')
                    unidad_medida = item.get('Unidad de medida')
                    id_proveedor = int(item.get('Proveedor'))
                    tipo = item.get('Tipo')
                    id_compra = item.get('Identificador')
                    

                    
                    proveedor_obj, created = Proveedores.objects.get_or_create(
                        id_proveedor=id_proveedor
                    )
                    
                    inventario_obj, created = Inventario.objects.get_or_create(
                        cod_inventario=codigo,
                        defaults={
                            'nombre': nombre,
                            'cantidad': cantidad,
                            'id_proveedor': proveedor_obj,
                            'tipo': tipo
                        }
                    )
                    
                    compra_obj, created = Compras.objects.get_or_create(
                        id_compra=id_compra,
                        defaults={
                            'estado': True,
                            'total_factura': 0
                        }
                    )
                    #Crear y guardar instancias en TransMp
                    nuevoSaldo = TransMp(
                        cod_inventario=inventario_obj,
                        nombre_mp=nombre,
                        cant_mp=cantidad,
                        costo_unitario=costo_unitario,
                        total_linea=total_linea,
                        fecha_ingreso=fecha_ingreso,
                        unidad_medida=unidad_medida,
                        id_proveedor=id_proveedor,
                        tipo=tipo,
                        id_compra=compra_obj
                    )
                    nuevoSaldo.save()
                
                        
                    
                    if not all([codigo, nombre, cantidad, costo_unitario, total_linea, fecha_ingreso, unidad_medida, id_proveedor, tipo, id_compra]):
                        logger.error("Faltan datos requeridos")
                    
                    logger.debug(f"Instancia de TransMp creada: {nuevoSaldo}")
                    
                except Exception as ve:
                    logger.error(f"Error al procesar el item: {item}, Error: {ve}")
                    
                    
                    
            for item in datos_tipo_pt:
                    codigo = int(item.get('Codigo'))
                    nombre = item.get('Nombre')
                    cantidad = int(item.get('Cantidad'))
                    costo_unitario = float(item.get('Costo Unitario', 0))
                    total_linea = float(item.get('Total', 0))         
                    fecha_ingreso= item.get('Fecha de ingreso')
                    unidad_medida = item.get('Unidad de medida')
                    id_proveedor = int(item.get('Proveedor'))
                    tipo = item.get('Tipo')
                    id_compra = item.get('Identificador')
                    
                    
                    proveedor_objt, created = Proveedores.objects.get_or_create(
                        id_proveedor=id_proveedor
                    )
                    
                    inventario_obj, created = Inventario.objects.get_or_create(
                        cod_inventario=codigo,
                        defaults={
                            'nombre': nombre,
                            'cantidad': cantidad,
                            'id_proveedor': proveedor_objt,
                            'tipo': tipo
                        }
                    )
                
                        
                    
            return JsonResponse({'status': 'Datos recibidos con éxito'})
        
        except json.JSONDecodeError:
            # Manejar errores de decodificación JSON
            return JsonResponse({'error': 'Error en el formato JSON'}, status=400)
        
        except Exception as e:
            # Manejar otros posibles errores
            return JsonResponse({'error': str(e)}, status=500)
    
    # Manejar el caso en que el método no sea POST
    return JsonResponse({'error': 'Método no permitido'}, status=405)


def verAverias(request):
    
    return render(request, 'verAverias.html')

def buscarAverias(request):
    fecha_inicio_str = request.GET.get('fecha_inicio')
    fecha_fin_str = request.GET.get('fecha_fin')

    # Convertir las fechas de cadena a objetos de fecha
    fecha_inicio = parse_date(fecha_inicio_str) if fecha_inicio_str else None
    fecha_fin = parse_date(fecha_fin_str) if fecha_fin_str else None

    # Filtrar facturas por rango de fechas
    averias_query = TransaccionAverias.objects.all()
    if fecha_inicio and fecha_fin:
        averias_query = averias_query.filter(fecha_averia__range=(fecha_inicio, fecha_fin))

    # Convertir y formatear las facturas filtradas
    averias_formateadas = []
    for averia in averias_query:
        id_averia = averia.id_averia.id_averia
        nombre_producto = get_object_or_404(Inventario, cod_inventario=averia.cod_inventario.cod_inventario)
        producto= nombre_producto.nombre
        costo= TransMp.objects.filter(cod_inventario=nombre_producto.cod_inventario).order_by('-fecha_ingreso').first()
        ultimo_precio = costo.costo_unitario if costo else 0.0 
        fecha = averia.fecha_averia
        cantidad = averia.cant_averia
        motivo = averia.motivo

        averia_formateada = {
            'id_averia': id_averia,
            'fecha_averia': fecha,
            'costo': float(ultimo_precio),
            'nombreProducto': producto,
            'cantidad': cantidad,
            'motivo': motivo
            
            
            
        }
        
        averias_formateadas.append(averia_formateada)

    return JsonResponse({'averias': averias_formateadas})



def imprimirInventario(request):
        # Crear un nuevo libro de trabajo de Excel
    wb = Workbook()
    ws = wb.active
    ws.title = 'inventario'

    # Definir los encabezados de las columnas
    columnas = ['ID', 'Nombre', 'bodega','fisico','total']
    ws.append(columnas)

    header_font = Font(bold=True, color='FFFFFF')  # Negrita y color blanco para texto
    header_fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type="solid")  # Fondo azul
    header_alignment = Alignment(horizontal="center", vertical="center")  # Alineación centrada

    for col in range(1, len(columnas) + 1):
        col_letter = get_column_letter(col)
        header_cell = ws[f'{col_letter}1']
        header_cell.font = header_font
        header_cell.fill = header_fill
        header_cell.alignment = header_alignment

    # Estilo para las celdas
    row_font = Font(name='Arial', size=10)  # Fuente Arial tamaño 10
    row_alignment = Alignment(horizontal="left", vertical="center")  # Alineación izquierda

    # Estilos para los bordes
    thin_border = Border(
        left=Side(border_style="thin"),
        right=Side(border_style="thin"),
        top=Side(border_style="thin"),
        bottom=Side(border_style="thin")
    )

    # Recuperar los datos de la base de datos
    productos = Inventario.objects.all().order_by('cod_inventario')
    
    # Crear rellenos para el formato condicional
    red_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")  # Relleno rojo claro
    green_fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")  # Relleno verde claro
    yellow_fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")  # Relleno amarillo claro

    for index, producto in enumerate(productos, start=2):  # Comienza desde la fila 2 (porque la fila 1 tiene los encabezados)
        # Escribir los datos en las columnas correspondientes
        ws[f'A{index}'] = producto.cod_inventario
        ws[f'B{index}'] = producto.nombre
        ws[f'C{index}'] = producto.cantidad  # Asumiendo que tienes un campo 'bodega'
        
        # Colocar la fórmula en la columna 'Total' (F)
        
        ws[f'E{index}'] = f'=D{index}-C{index}'  # Ejemplo de fórmula (fisico * bodega)
        
        total_cell = ws[f'E{index}']
        total_value = f'=D{index}-C{index}'  # Fórmula de la celda
        
                # Verificar el valor de la celda "Total" y aplicar el formato adecuado
        ws[f'E{index}'].value = total_value

        

            # Aplicar estilos a las celdas de cada fila
        for col in range(1, 6):  # Para las columnas A a E
            cell = ws.cell(row=index, column=col)
            cell.font = row_font
            cell.alignment = row_alignment
            cell.border = thin_border
    # Ajustar el ancho de las columnas
    for col in range(1, len(columnas) + 1):
        column_letter = get_column_letter(col)
        ws.column_dimensions[column_letter].width = 20
        
        
    # Obtener la fecha y hora actuales para el nombre del archivo
    current_datetime = datetime.now().strftime('%Y-%m-%d_%H-%M-%S')  # Formato: '2025-01-15_12-30-45'
    filename = f'Inventario_{current_datetime}.xlsx'  # Crear el nombre de archivo    

    # Crear una respuesta HTTP para descargar el archivo
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename={filename}'

    # Guardar el libro de trabajo en el objeto HttpResponse
    wb.save(response)
    return response



def CargarDatos(request):
    return render(request, 'cargarInventario.html')


def ProcesarInventario(request):
    if request.method == 'POST':
        print('entramos a back')
        # Obtener los datos del formulario
        excel_file = request.FILES.get('file')
        if not excel_file:
            return JsonResponse({'error': 'No se ha subido ningún archivo'}, status=400)
        try:
            # Leer el archivo Excel con pandas
            df = pd.read_excel(excel_file, engine='openpyxl')
            
            # Reemplazar NaN en la columna 'bodega' con 0
            df['fisico'] = df['fisico'].fillna(0)
            df['total']= df['fisico'] - df['bodega']
            
            print('df: ', df)
            
            
            
            # Verificar que el archivo tiene la columna E (usaremos el índice 4 para la columna E)
            if 'total' not in df.columns:
                return JsonResponse({'error': 'La columna total no está presente en el archivo'}, status=400)
            
            # Filtrar filas donde los valores en la columna E sean mayores que 0 o menores que 0
            df_filtrado = df[(df['total'] > 0) | (df['total'] < 0)]
            
            # Verificar que las columnas requeridas existen en el archivo
            required_columns = ['ID', 'Nombre', 'bodega', 'fisico', 'total']
            for col in required_columns:
                if col not in df.columns:
                    return JsonResponse({'error': f'Falta la columna {col} en el archivo'}, status=400)

            # Crear una lista de diccionarios con los datos filtrados
            datos = df_filtrado[required_columns].to_dict(orient='records')
            
            # Enumerar los registros agregando un campo 'id' dinámico
            for index, item in enumerate(datos, start=1):
                item['id'] = index  # Agregar el campo 'id' con la numeración
            
            # Retornar los datos filtrados al frontend (puedes estructurarlo como una tabla o lista)
            return JsonResponse({'data': datos}, status=200)
        
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)
        
        
        
def EditarNombreInven(request):
    if request.method == 'POST':
        try:
            cod_inventario = request.POST.get('id')
            nombre = request.POST.get('new_name')
            
            Producto= Inventario.objects.get(cod_inventario = cod_inventario)
            
            Producto.nombre = nombre
            Producto.save()
            return JsonResponse({'message': 'Nombre actualizado correctamente'}, status=200)
        except Inventario.DoesNotExist:
            return JsonResponse({'error': 'No existe el inventario'}, status=404)
        except Exception as e:
            return JsonResponse({'sucess':False, 'error': str(e)}, status=500)
        
                            
                            
                            
                            
                            
def GenerarAjusteInv(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            descripcion = 'Ajuste de Inventario Fisico'
            fecha = datetime.now().date()
            print('fecha: ', fecha , 'datos:', data)
            ultimo_ajuste = None
            ultimo_ajuste = Ajustes.objects.aggregate(Max('id_ajuste'))['id_ajuste__max']
            print(ultimo_ajuste)
            if ultimo_ajuste is not None:
                nuevo_ajuste = ultimo_ajuste + 1
                print('ajuste', nuevo_ajuste)
            else:
                nuevo_ajuste = 10001
                print(nuevo_ajuste)
            
            
            ajusteN, _ = Ajustes.objects.get_or_create(id_ajuste = nuevo_ajuste)
            for producto in data:
                
                cod_inventario = producto['cod_inventario']
                nombre = producto['nombre']
                cantidad = round(float(producto['total']),2)
                print('codigo: ', cod_inventario, 'cantidad: ', cantidad)
                
                
                
                inventario, _ = Inventario.objects.get_or_create(cod_inventario= cod_inventario)
                nueva_cantidad = F('cantidad') + cantidad
                
                nuevo_ajuste = TransaccionAjuste.objects.create(
                fecha_ajuste = fecha,
                cant_ajuste = cantidad,
                descripcion = descripcion,
                id_ajuste = ajusteN,
                cod_inventario = inventario
                )
                nuevo_ajuste.save();
                # Aquí puedes procesar y guardar los datos adicionales si es necesario
                Inventario.objects.filter(cod_inventario=cod_inventario).update(cantidad=nueva_cantidad)

            # Retornar una respuesta de éxito
            return JsonResponse({'status': 'success','mensaje': 'Ajuste generado correctamente'}, status=200)

        except json.JSONDecodeError:
            return JsonResponse({'error': 'Error al procesar los datos del JSON'}, status=400)
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)
            
            
        